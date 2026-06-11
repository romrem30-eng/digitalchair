from decimal import Decimal
from decimal import InvalidOperation
from io import BytesIO

from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl import load_workbook

from apps.core.models import StudentGroup
from apps.core.models import Subject
from apps.core.models import Teacher
from apps.core.models import WorkloadPlan
from apps.workload.models import WorkloadAssignment


IMPORT_DEFINITIONS = {
    'teachers': {
        'title': 'Преподаватели',
        'template_filename': 'teachers_template.xlsx',
        'headers': (
            'ФИО',
            'ученая степень',
            'звание',
            'ставка',
            'максимальная нагрузка',
            'контакты',
        ),
        'sample_row': (
            'Иванов Иван Иванович',
            'к.т.н.',
            'доцент',
            '1.0',
            '900',
            '+79990001122',
        ),
    },
    'subjects': {
        'title': 'Дисциплины',
        'template_filename': 'subjects_template.xlsx',
        'headers': (
            'название',
            'код',
            'семестр',
            'трудоемкость',
            'форма контроля',
        ),
        'sample_row': (
            'Программирование',
            'INF-101',
            '1',
            '144',
            'экзамен',
        ),
    },
    'groups': {
        'title': 'Учебные группы',
        'template_filename': 'groups_template.xlsx',
        'headers': (
            'название',
            'курс',
            'направление подготовки',
        ),
        'sample_row': (
            'ИВТ-101',
            '1',
            'Информатика и вычислительная техника',
        ),
    },
    'workload': {
        'title': 'Учебный план нагрузки',
        'template_filename': 'workload_template.xlsx',
        'headers': (
            'дисциплина',
            'часы',
            'семестр',
        ),
        'sample_row': (
            'Программирование',
            '144',
            '1',
        ),
    },
}


POSITION_MAP = {
    'assistant': 'assistant',
    'ассистент': 'assistant',
    'senior_teacher': 'senior_teacher',
    'старший преподаватель': 'senior_teacher',
    'docent': 'docent',
    'доцент': 'docent',
    'professor': 'professor',
    'профессор': 'professor',
}


CONTROL_TYPE_MAP = {
    'exam': 'exam',
    'экзамен': 'exam',
    'test': 'test',
    'зачет': 'test',
    'зачёт': 'test',
    'coursework': 'coursework',
    'курсовая работа': 'coursework',
}


def normalize_string(value):

    if value is None:

        return ''

    return str(value).strip()


def normalize_header(value):

    return normalize_string(value).lower()


def parse_positive_int(value, field_label):

    raw_value = normalize_string(value)

    if not raw_value:

        raise ValueError(f'Поле "{field_label}" обязательно.')

    try:

        number = int(float(raw_value))

    except (TypeError, ValueError):

        raise ValueError(
            f'Поле "{field_label}" должно быть целым числом.'
        )

    if number < 0:

        raise ValueError(
            f'Поле "{field_label}" не может быть отрицательным.'
        )

    return number


def parse_decimal_value(value, field_label):

    raw_value = normalize_string(value).replace(',', '.')

    if not raw_value:

        raise ValueError(f'Поле "{field_label}" обязательно.')

    try:

        number = Decimal(raw_value)

    except InvalidOperation:

        raise ValueError(
            f'Поле "{field_label}" должно быть числом.'
        )

    if number < 0:

        raise ValueError(
            f'Поле "{field_label}" не может быть отрицательным.'
        )

    return number


def parse_teacher_position(value):

    normalized_value = normalize_header(value)

    if normalized_value not in POSITION_MAP:

        raise ValueError(
            'Поле "звание" должно содержать одно из значений: '
            'ассистент, старший преподаватель, доцент, профессор.'
        )

    return POSITION_MAP[normalized_value]


def parse_control_type(value):

    normalized_value = normalize_header(value)

    if normalized_value not in CONTROL_TYPE_MAP:

        raise ValueError(
            'Поле "форма контроля" должно содержать одно из значений: '
            'экзамен, зачет, курсовая работа.'
        )

    return CONTROL_TYPE_MAP[normalized_value]


def build_template_workbook(import_type):

    definition = IMPORT_DEFINITIONS[import_type]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = definition['title']

    worksheet.append(definition['headers'])
    worksheet.append(definition['sample_row'])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def get_template_filename(import_type):

    return IMPORT_DEFINITIONS[import_type]['template_filename']


def get_import_title(import_type):

    return IMPORT_DEFINITIONS[import_type]['title']


def read_workbook_rows(uploaded_file, import_type):

    uploaded_file.seek(0)

    workbook = load_workbook(
        uploaded_file,
        data_only=True
    )

    worksheet = workbook.active
    expected_headers = [
        normalize_header(header)
        for header in IMPORT_DEFINITIONS[import_type]['headers']
    ]

    header_row = next(
        worksheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True
        )
    )

    actual_headers = [
        normalize_header(value)
        for value in header_row[:len(expected_headers)]
    ]

    if actual_headers != expected_headers:

        return None, [
            'Структура файла не соответствует шаблону.'
        ]

    rows = []

    for row_index, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        values = row[:len(expected_headers)]

        if all(
            normalize_string(value) == ''
            for value in values
        ):

            continue

        rows.append((row_index, values))

    return rows, []


def build_import_preview(import_type, uploaded_file, plan=None):

    rows, header_errors = read_workbook_rows(
        uploaded_file,
        import_type
    )

    preview = {
        'import_type': import_type,
        'title': get_import_title(import_type),
        'total_rows': 0,
        'new_count': 0,
        'update_count': 0,
        'error_count': 0,
        'errors': [],
        'prepared_rows': [],
        'selected_plan_id': plan.id if plan else None,
        'selected_plan_label': str(plan) if plan else '',
        'summary': '',
    }

    if header_errors:

        preview['errors'] = header_errors
        preview['error_count'] = len(header_errors)
        return preview

    preview['total_rows'] = len(rows)

    if import_type == 'teachers':

        populate_teacher_preview(
            preview,
            rows
        )

    elif import_type == 'subjects':

        populate_subject_preview(
            preview,
            rows
        )

    elif import_type == 'groups':

        populate_group_preview(
            preview,
            rows
        )

    elif import_type == 'workload':

        populate_workload_preview(
            preview,
            rows,
            plan
        )

    preview['error_count'] = len(preview['errors'])
    preview['summary'] = (
        f'Строк: {preview["total_rows"]}, '
        f'новых: {preview["new_count"]}, '
        f'обновляемых: {preview["update_count"]}, '
        f'ошибок: {preview["error_count"]}.'
    )

    return preview


def populate_teacher_preview(preview, rows):

    seen_names = set()

    for row_index, values in rows:

        try:

            full_name = normalize_string(values[0])
            academic_degree = normalize_string(values[1])
            position = parse_teacher_position(values[2])
            rate = parse_decimal_value(
                values[3],
                'ставка'
            )
            max_hours = parse_positive_int(
                values[4],
                'максимальная нагрузка'
            )
            contacts = normalize_string(values[5])

            if not full_name:

                raise ValueError(
                    'Поле "ФИО" обязательно.'
                )

            if contacts and len(contacts) > 100:

                raise ValueError(
                    'Поле "контакты" не должно превышать 100 символов.'
                )

            normalized_name = full_name.lower()

            if normalized_name in seen_names:

                raise ValueError(
                    'В файле найден дубликат преподавателя.'
                )

            seen_names.add(normalized_name)

            teacher = Teacher.objects.filter(
                full_name=full_name
            ).select_related('user').first()

            preview['prepared_rows'].append({
                'mode': 'update' if teacher else 'create',
                'full_name': full_name,
                'academic_degree': academic_degree,
                'position': position,
                'rate': str(rate),
                'max_hours': max_hours,
                'contacts': contacts,
            })

            if teacher:

                preview['update_count'] += 1

            else:

                preview['new_count'] += 1

        except ValueError as error:

            preview['errors'].append(
                f'Строка {row_index}: {error}'
            )


def populate_subject_preview(preview, rows):

    seen_keys = set()

    for row_index, values in rows:

        try:

            name = normalize_string(values[0])
            code = normalize_string(values[1])
            semester = parse_positive_int(
                values[2],
                'семестр'
            )
            hours = parse_positive_int(
                values[3],
                'трудоемкость'
            )
            control_type = parse_control_type(values[4])

            if not name:

                raise ValueError(
                    'Поле "название" обязательно.'
                )

            if not code:

                raise ValueError(
                    'Поле "код" обязательно.'
                )

            key = (name.lower(), semester)

            if key in seen_keys:

                raise ValueError(
                    'В файле найден дубликат дисциплины.'
                )

            seen_keys.add(key)

            subject = Subject.objects.filter(
                name=name,
                semester=semester
            ).first()

            preview['prepared_rows'].append({
                'mode': 'update' if subject else 'create',
                'name': name,
                'code': code,
                'semester': semester,
                'hours': hours,
                'control_type': control_type,
            })

            if subject:

                preview['update_count'] += 1

            else:

                preview['new_count'] += 1

        except ValueError as error:

            preview['errors'].append(
                f'Строка {row_index}: {error}'
            )


def populate_group_preview(preview, rows):

    seen_names = set()

    for row_index, values in rows:

        try:

            name = normalize_string(values[0])
            course = parse_positive_int(
                values[1],
                'курс'
            )
            direction = normalize_string(values[2])

            if not name:

                raise ValueError(
                    'Поле "название" обязательно.'
                )

            if not direction:

                raise ValueError(
                    'Поле "направление подготовки" обязательно.'
                )

            normalized_name = name.lower()

            if normalized_name in seen_names:

                raise ValueError(
                    'В файле найден дубликат группы.'
                )

            seen_names.add(normalized_name)

            group = StudentGroup.objects.filter(
                name=name
            ).first()

            preview['prepared_rows'].append({
                'mode': 'update' if group else 'create',
                'name': name,
                'course': course,
                'direction': direction,
            })

            if group:

                preview['update_count'] += 1

            else:

                preview['new_count'] += 1

        except ValueError as error:

            preview['errors'].append(
                f'Строка {row_index}: {error}'
            )


def populate_workload_preview(preview, rows, plan):

    if plan is None:

        preview['errors'].append(
            'Не выбран план нагрузки для импорта.'
        )
        return

    if plan.status == WorkloadPlan.Statuses.APPROVED:

        preview['errors'].append(
            'Нельзя импортировать данные в утвержденный план нагрузки.'
        )

    if WorkloadAssignment.objects.filter(plan=plan).exists():

        preview['errors'].append(
            'Нельзя импортировать учебный план в план, по которому уже есть распределение нагрузки.'
        )

    seen_keys = set()
    total_hours = 0

    for row_index, values in rows:

        try:

            name = normalize_string(values[0])
            hours = parse_positive_int(
                values[1],
                'часы'
            )
            semester = parse_positive_int(
                values[2],
                'семестр'
            )

            if not name:

                raise ValueError(
                    'Поле "дисциплина" обязательно.'
                )

            key = (name.lower(), semester)

            if key in seen_keys:

                raise ValueError(
                    'В файле найден дубликат дисциплины.'
                )

            seen_keys.add(key)
            total_hours += hours

            subject = Subject.objects.filter(
                name=name,
                semester=semester
            ).first()

            preview['prepared_rows'].append({
                'mode': 'update' if subject else 'create',
                'name': name,
                'hours': hours,
                'semester': semester,
            })

            if subject:

                preview['update_count'] += 1

            else:

                preview['new_count'] += 1

        except ValueError as error:

            preview['errors'].append(
                f'Строка {row_index}: {error}'
            )

    preview['calculated_total_hours'] = total_hours


def generate_unique_teacher_email(full_name):

    User = get_user_model()
    base_slug = slugify(full_name) or 'teacher'
    suffix = 1

    while True:

        email = f'{base_slug}-{suffix}@digitalchair.local'

        if not User.objects.filter(
            email=email
        ).exists():

            return email

        suffix += 1


def assign_contacts_to_user(user, contacts):

    if not contacts:

        return []

    if len(contacts) <= 20:

        user.phone = contacts
        user.telegram_id = ''

        return ['phone', 'telegram_id']

    user.phone = ''
    user.telegram_id = contacts

    return ['phone', 'telegram_id']


@transaction.atomic
def apply_import_preview(preview):

    import_type = preview['import_type']

    if import_type == 'teachers':

        return apply_teacher_rows(
            preview['prepared_rows']
        )

    if import_type == 'subjects':

        return apply_subject_rows(
            preview['prepared_rows']
        )

    if import_type == 'groups':

        return apply_group_rows(
            preview['prepared_rows']
        )

    return apply_workload_rows(preview)


def apply_teacher_rows(rows):

    User = get_user_model()
    applied_count = 0

    for row in rows:

        teacher = Teacher.objects.filter(
            full_name=row['full_name']
        ).select_related('user').first()

        if teacher is None:

            email = generate_unique_teacher_email(
                row['full_name']
            )

            user = User.objects.create_user(
                email=email,
                password=None,
                role='TEACHER'
            )

            teacher = Teacher.objects.create(
                user=user,
                full_name=row['full_name'],
                position=row['position'],
                academic_degree=row['academic_degree'],
                rate=Decimal(row['rate']),
                max_hours=row['max_hours'],
            )

        else:

            teacher.full_name = row['full_name']
            teacher.position = row['position']
            teacher.academic_degree = row['academic_degree']
            teacher.rate = Decimal(row['rate'])
            teacher.max_hours = row['max_hours']
            teacher.save(
                update_fields=[
                    'full_name',
                    'position',
                    'academic_degree',
                    'rate',
                    'max_hours',
                ]
            )

        user_fields = assign_contacts_to_user(
            teacher.user,
            row['contacts']
        )

        if user_fields:

            teacher.user.save(
                update_fields=user_fields
            )

        applied_count += 1

    return applied_count


def apply_subject_rows(rows):

    applied_count = 0

    for row in rows:

        subject, created = Subject.objects.get_or_create(
            name=row['name'],
            semester=row['semester'],
            defaults={
                'hours': row['hours'],
                'control_type': row['control_type'],
            }
        )

        if not created:

            subject.hours = row['hours']
            subject.control_type = row['control_type']
            subject.save(
                update_fields=[
                    'hours',
                    'control_type'
                ]
            )

        applied_count += 1

    return applied_count


def apply_group_rows(rows):

    applied_count = 0

    for row in rows:

        group, created = StudentGroup.objects.get_or_create(
            name=row['name'],
            defaults={
                'course': row['course'],
                'direction': row['direction'],
            }
        )

        if not created:

            group.course = row['course']
            group.direction = row['direction']
            group.save(
                update_fields=[
                    'course',
                    'direction'
                ]
            )

        applied_count += 1

    return applied_count


def apply_workload_rows(preview):

    applied_count = 0
    plan = WorkloadPlan.objects.get(
        pk=preview['selected_plan_id']
    )

    for row in preview['prepared_rows']:

        subject = Subject.objects.filter(
            name=row['name'],
            semester=row['semester']
        ).first()

        if subject is None:

            Subject.objects.create(
                name=row['name'],
                hours=row['hours'],
                semester=row['semester'],
                control_type='test'
            )

        else:

            subject.hours = row['hours']
            subject.semester = row['semester']
            subject.save(
                update_fields=[
                    'hours',
                    'semester'
                ]
            )

        applied_count += 1

    plan.total_hours = preview.get(
        'calculated_total_hours',
        plan.total_hours
    )
    plan.save(
        update_fields=[
            'total_hours'
        ]
    )

    return applied_count
